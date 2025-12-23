#!/usr/bin/env python3
"""
Political Tendency Analyzer - Web Server
Scrapes news URLs and determines political tendency based on keywords
"""

import json
import os
import re
import shutil
import sys
import atexit
import signal
from flask import Flask, render_template, request, jsonify, send_file
from werkzeug.utils import secure_filename
import requests
from bs4 import BeautifulSoup
from urllib.parse import urlparse

# Try to import pyserial for Arduino communication
try:
    import serial
    import serial.tools.list_ports
    SERIAL_SUPPORT = True
except ImportError:
    SERIAL_SUPPORT = False
    print("Warning: pyserial not installed. Arduino serial communication disabled.")
    print("Install with: pip install pyserial")

# Try to import openpyxl for Excel support
try:
    from openpyxl import load_workbook
    EXCEL_SUPPORT = True
except ImportError:
    EXCEL_SUPPORT = False
    print("Warning: openpyxl not installed. Excel file support disabled. Install with: pip install openpyxl")

# Handle PyInstaller bundle path
VERSION = "1.3.0"

app = Flask(__name__)
if getattr(sys, 'frozen', False):
    # Running as compiled executable
    base_path = sys._MEIPASS
    template_folder = os.path.join(base_path, 'templates')
    # ... (verification logic) ...
    app = Flask(__name__, template_folder=template_folder)
else:
    # Running as script
    app = Flask(__name__)

# Global error handler for 500 errors
@app.errorhandler(500)
def internal_error(error):
    import traceback
    err_trace = traceback.format_exc()
    print(f"[DEBUG] 500 ERROR CAUGHT: {error}", flush=True)
    print(err_trace, flush=True)
    return jsonify({'error': '서버 내부 오류가 발생했습니다.', 'details': str(error)}), 500

@app.errorhandler(Exception)
def handle_exception(e):
    import traceback
    err_trace = traceback.format_exc()
    print(f"[DEBUG] UNHANDLED EXCEPTION: {e}", flush=True)
    print(err_trace, flush=True)
    return jsonify({'error': '예기치 못한 오류가 발생했습니다.', 'details': str(e)}), 500

# Configure logging to ensure debug output is visible
import logging
logging.basicConfig(level=logging.DEBUG)
app.logger.setLevel(logging.DEBUG)
app.logger.info("Flask app initialized")

# Global serial connection
arduino_serial = None
serial_port = None

def init_arduino_serial(port=None, baudrate=9600):
    """Initialize serial connection to Arduino"""
    global arduino_serial, serial_port
    
    if not SERIAL_SUPPORT:
        return False
    
    try:
        # If port not specified, try to auto-detect
        if port is None:
            ports = serial.tools.list_ports.comports()
            # Try to find Arduino ports (works on Windows, Mac, Linux)
            arduino_ports = []
            for p in ports:
                desc_lower = p.description.lower() if p.description else ''
                device_lower = p.device.lower() if p.device else ''
                # Check for Arduino indicators (works across platforms)
                if ('arduino' in desc_lower or 
                    'usb serial' in desc_lower or
                    'ch340' in desc_lower or  # Common Arduino clone chip
                    'cp210' in desc_lower or   # Another common chip
                    'ftdi' in desc_lower):     # FTDI chip
                    arduino_ports.append(p.device)
            
            if arduino_ports:
                port = arduino_ports[0]
                print(f"Auto-detected Arduino port: {port}")
            else:
                # If no Arduino-specific port found, list all available ports
                print("No Arduino-specific port detected. Available ports:")
                for p in ports:
                    print(f"  - {p.device}: {p.description or 'No description'}")
                # On Windows, COM ports are common; on Mac/Linux, /dev/tty* or /dev/cu*
                # User can manually select from the web interface
                return False
        
        # Close existing connection if any
        if arduino_serial and arduino_serial.is_open:
            try:
                arduino_serial.close()
            except:
                pass
        
        arduino_serial = serial.Serial(port, baudrate, timeout=1, write_timeout=1)
        serial_port = port
        # Clear buffers
        arduino_serial.reset_input_buffer()
        arduino_serial.reset_output_buffer()
        print(f"Arduino connected on {port} at {baudrate} baud")
        
        # Wait a moment for Arduino to be ready, then send neutral value (500)
        import time
        time.sleep(0.5)  # Give Arduino time to initialize
        send_to_arduino(500)
        
        return True
    except serial.SerialException as e:
        print(f"Serial error connecting to Arduino: {str(e)}")
        arduino_serial = None
        return False
    except Exception as e:
        print(f"Error connecting to Arduino: {str(e)}")
        arduino_serial = None
        return False

def send_to_arduino(value):
    """Send value to Arduino via serial with error recovery"""
    global arduino_serial, serial_port
    
    if not SERIAL_SUPPORT:
        return False
    
    # Check if connection is valid
    is_connected = False
    try:
        if arduino_serial is not None:
            is_connected = arduino_serial.is_open
    except (AttributeError, ValueError, OSError) as e:
        print(f"Serial port in invalid state: {e}")
        arduino_serial = None
        is_connected = False
    
    if not is_connected:
        print("Arduino serial connection not available, attempting to reconnect...")
        # Try to reconnect if we have a port
        if serial_port:
            try:
                arduino_serial = serial.Serial(serial_port, 9600, timeout=1, write_timeout=1)
                arduino_serial.reset_input_buffer()
                arduino_serial.reset_output_buffer()
                print(f"Reconnected to Arduino on {serial_port}")
                # Continue to send the requested value after reconnection
            except Exception as e:
                print(f"Failed to reconnect to Arduino: {e}")
                arduino_serial = None
                return False
        else:
            return False
    
    try:
        # Clear any pending input
        arduino_serial.reset_input_buffer()
        
        # Send value as integer string with newline (compatible with Serial.parseInt())
        # Format: "750\n" - Arduino's Serial.parseInt() will read this correctly
        message = f"{value}\n"
        
        # Send the command 10 times with 0.1 second intervals to ensure it's received
        import time
        for i in range(10):
            arduino_serial.write(message.encode('utf-8'))
            arduino_serial.flush()
            print(f"[{i+1}/10] Sending to Arduino: {value}")
            if i < 9:  # Don't delay after the last send
                time.sleep(0.1)
        
        print(f"Completed sending {value} to Arduino (sent 10 times)")
        return True
    except serial.SerialException as e:
        print(f"Serial error sending to Arduino: {str(e)}")
        # Close and mark as disconnected
        try:
            arduino_serial.close()
        except:
            pass
        arduino_serial = None
        return False
    except Exception as e:
        print(f"Error sending to Arduino: {str(e)}")
        return False

def cleanup_arduino():
    """Cleanup function to disconnect Arduino on program exit"""
    global arduino_serial, serial_port
    
    if SERIAL_SUPPORT and arduino_serial is not None:
        try:
            if arduino_serial.is_open:
                arduino_serial.close()
                print(f"Arduino disconnected from {serial_port}")
            arduino_serial = None
            serial_port = None
        except Exception as e:
            print(f"Error disconnecting Arduino during cleanup: {e}")

# Register cleanup function to run on exit
atexit.register(cleanup_arduino)

# Register signal handlers for graceful shutdown
def signal_handler(signum, frame):
    """Handle interrupt signals (Ctrl+C, etc.)"""
    print("\n\nShutting down server...")
    cleanup_arduino()
    sys.exit(0)

# Register signal handlers (SIGINT works on both Windows and Unix)
signal.signal(signal.SIGINT, signal_handler)
# SIGTERM is Unix-only, so wrap in try-except for Windows compatibility
try:
    signal.signal(signal.SIGTERM, signal_handler)
except AttributeError:
    # SIGTERM not available on Windows
    pass

# Load keywords from Excel or JSON files
def load_keywords():
    """Load progressive and conservative keywords from Excel or JSON files"""
    # Handle PyInstaller bundle path
    if getattr(sys, 'frozen', False):
        # Running as compiled executable
        # First check executable directory (where users can place files)
        exe_dir = os.path.dirname(sys.executable)
        bundle_dir = sys._MEIPASS  # Bundled files location
    else:
        # Running as script
        exe_dir = os.path.dirname(__file__)
        bundle_dir = exe_dir
    
    def find_file(filename):
        """Find file in executable directory first, then bundle directory"""
        exe_path = os.path.join(exe_dir, filename)
        if os.path.exists(exe_path):
            return exe_path
        bundle_path = os.path.join(bundle_dir, filename)
        if os.path.exists(bundle_path):
            return bundle_path
        return None
    
    # Excel file paths
    progressive_excel = find_file('progressive.xlsx')
    conservative_excel = find_file('conservative.xlsx')
    keywords_excel = find_file('keywords.xlsx')  # Single file with multiple sheets
    
    # JSON file paths
    progressive_json = find_file('progressive.json')
    conservative_json = find_file('conservative.json')
    
    progressive_keywords = []
    conservative_keywords = []
    
    # Try to load from Excel files first
    if EXCEL_SUPPORT:
        # Check for single Excel file with multiple sheets
        if keywords_excel and os.path.exists(keywords_excel):
            try:
                wb = load_workbook(keywords_excel, read_only=True)
                sheet_names = wb.sheetnames
                
                # Look for progressive sheet (check both English and Korean names)
                progressive_sheet = None
                for sheet_name in sheet_names:
                    if 'progressive' in sheet_name.lower() or '진보' in sheet_name:
                        progressive_sheet = sheet_name
                        break
                
                if progressive_sheet:
                    ws = wb[progressive_sheet]
                    progressive_keywords = []
                    # Read from all columns, not just the first one
                    max_col = ws.max_column
                    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=max_col, values_only=True):
                        for cell_value in row:
                            if cell_value:
                                # Split cell value into multiple keywords if it contains delimiters
                                cell_keywords = split_keywords(cell_value)
                                for kw in cell_keywords:
                                    # Skip header rows (check both English and Korean)
                                    kw_lower = kw.lower()
                                    header_keywords = ['keyword', 'keywords', 'term', 'terms', 'progressive', 
                                                      '키워드', '단어', '용어', '진보', '진보진영']
                                    if kw_lower not in header_keywords and kw not in header_keywords:
                                        progressive_keywords.append(kw)
                    progressive_keywords = [kw for kw in progressive_keywords if kw]
                
                # Look for conservative sheet (check both English and Korean names)
                conservative_sheet = None
                for sheet_name in sheet_names:
                    if 'conservative' in sheet_name.lower() or '보수' in sheet_name:
                        conservative_sheet = sheet_name
                        break
                
                if conservative_sheet:
                    ws = wb[conservative_sheet]
                    conservative_keywords = []
                    # Read from all columns, not just the first one
                    max_col = ws.max_column
                    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=max_col, values_only=True):
                        for cell_value in row:
                            if cell_value:
                                # Split cell value into multiple keywords if it contains delimiters
                                cell_keywords = split_keywords(cell_value)
                                for kw in cell_keywords:
                                    # Skip header rows (check both English and Korean)
                                    kw_lower = kw.lower()
                                    header_keywords = ['keyword', 'keywords', 'term', 'terms', 'conservative',
                                                      '키워드', '단어', '용어', '보수', '보수진영']
                                    if kw_lower not in header_keywords and kw not in header_keywords:
                                        conservative_keywords.append(kw)
                    conservative_keywords = [kw for kw in conservative_keywords if kw]
                
                wb.close()
                
                if progressive_keywords or conservative_keywords:
                    print(f"Loaded keywords from {keywords_excel}")
            except Exception as e:
                print(f"Warning: Error reading {keywords_excel}: {str(e)}")
        
        # Check for separate Excel files
        if not progressive_keywords and progressive_excel and os.path.exists(progressive_excel):
            try:
                wb = load_workbook(progressive_excel, read_only=True)
                ws = wb.active
                progressive_keywords = []
                # Read from all columns, not just the first one
                max_col = ws.max_column
                for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=max_col, values_only=True):
                    for cell_value in row:
                        if cell_value:
                            # Split cell value into multiple keywords if it contains delimiters
                            cell_keywords = split_keywords(cell_value)
                            for kw in cell_keywords:
                                # Skip header rows (check both English and Korean)
                                kw_lower = kw.lower()
                                header_keywords = ['keyword', 'keywords', 'term', 'terms', 'progressive',
                                                  '키워드', '단어', '용어', '진보', '진보진영']
                                if kw_lower not in header_keywords and kw not in header_keywords:
                                    progressive_keywords.append(kw)
                progressive_keywords = [kw for kw in progressive_keywords if kw]
                wb.close()
                print(f"Loaded {len(progressive_keywords)} progressive keywords from {progressive_excel}")
            except Exception as e:
                print(f"Warning: Error reading {progressive_excel}: {str(e)}")
        
        if not conservative_keywords and conservative_excel and os.path.exists(conservative_excel):
            try:
                wb = load_workbook(conservative_excel, read_only=True)
                ws = wb.active
                conservative_keywords = []
                # Read from all columns, not just the first one
                max_col = ws.max_column
                for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=max_col, values_only=True):
                    for cell_value in row:
                        if cell_value:
                            # Split cell value into multiple keywords if it contains delimiters
                            cell_keywords = split_keywords(cell_value)
                            for kw in cell_keywords:
                                # Skip header rows (check both English and Korean)
                                kw_lower = kw.lower()
                                header_keywords = ['keyword', 'keywords', 'term', 'terms', 'conservative',
                                                  '키워드', '단어', '용어', '보수', '보수진영']
                                if kw_lower not in header_keywords and kw not in header_keywords:
                                    conservative_keywords.append(kw)
                conservative_keywords = [kw for kw in conservative_keywords if kw]
                wb.close()
                print(f"Loaded {len(conservative_keywords)} conservative keywords from {conservative_excel}")
            except Exception as e:
                print(f"Warning: Error reading {conservative_excel}: {str(e)}")
    
    # Fall back to JSON files if Excel files not found or failed
    if not progressive_keywords and progressive_json and os.path.exists(progressive_json):
        try:
            with open(progressive_json, 'r', encoding='utf-8') as f:
                progressive_keywords = json.load(f)
            print(f"Loaded {len(progressive_keywords)} progressive keywords from {progressive_json}")
        except Exception as e:
            print(f"Warning: Error reading {progressive_json}: {str(e)}")
    
    if not conservative_keywords and conservative_json and os.path.exists(conservative_json):
        try:
            with open(conservative_json, 'r', encoding='utf-8') as f:
                conservative_keywords = json.load(f)
            print(f"Loaded {len(conservative_keywords)} conservative keywords from {conservative_json}")
        except Exception as e:
            print(f"Warning: Error reading {conservative_json}: {str(e)}")
    
    return progressive_keywords, conservative_keywords

def scrape_article(url):
    """Scrape article content from a given URL with extreme resilience"""
    import sys
    import urllib.request
    import urllib.error
    import ssl
    import gzip
    import zlib
    
    print(f"[DEBUG] ===== scrape_article START: {url} =====", flush=True)
    
    try:
        # SSL Context
        ssl_context = ssl.create_default_context()
        ssl_context.check_hostname = False
        ssl_context.verify_mode = ssl.CERT_NONE
        
        # Headers
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
            'Accept-Language': 'ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7',
            'Accept-Encoding': 'identity', # Explicitly request uncompressed
            'Connection': 'keep-alive'
        }
        
        req = urllib.request.Request(url, headers=headers)
        
        print(f"[DEBUG] Step 1: Requesting URL...", flush=True)
        with urllib.request.urlopen(req, timeout=15, context=ssl_context) as response:
            raw_data = response.read()
            info = response.info()
            content_encoding = info.get('Content-Encoding', '').lower()
            content_type = info.get('Content-Type', '')
            
            print(f"[DEBUG] Step 1 SUCCESS: Received {len(raw_data)} bytes", flush=True)
            print(f"[DEBUG] Content-Type: {content_type}, Encoding: {content_encoding}", flush=True)
            
            # 2. Decompression (Only if absolutely necessary)
            processed_data = raw_data
            
            # Check for GZIP magic numbers (1f 8b)
            if raw_data.startswith(b'\x1f\x8b'):
                print(f"[DEBUG] GZIP signature detected, decompressing...", flush=True)
                try:
                    processed_data = gzip.decompress(raw_data)
                    print(f"[DEBUG] GZIP decompression success: {len(processed_data)} bytes", flush=True)
                except Exception as e:
                    print(f"[DEBUG] GZIP decompression FAILED: {e}", flush=True)
            
            # If not gzip, but header says deflate or we suspect it
            elif content_encoding == 'deflate' or content_encoding == 'gzip':
                print(f"[DEBUG] Content-Encoding header '{content_encoding}' found, attempting decompression...", flush=True)
                try:
                    # Try standard zlib
                    processed_data = zlib.decompress(raw_data)
                    print(f"[DEBUG] Zlib decompression success: {len(processed_data)} bytes", flush=True)
                except Exception:
                    try:
                        # Try raw deflate
                        processed_data = zlib.decompress(raw_data, -zlib.MAX_WBITS)
                        print(f"[DEBUG] Raw Deflate decompression success: {len(processed_data)} bytes", flush=True)
                    except Exception as e:
                        print(f"[DEBUG] All decompression FAILED: {e}", flush=True)
            
            # 3. Decoding
            html_content = None
            
            # Detect charset from header
            charset = 'utf-8'
            if 'charset=' in content_type.lower():
                try:
                    charset = content_type.lower().split('charset=')[-1].split(';')[0].strip()
                except: pass
            
            encodings_to_try = [charset, 'utf-8', 'cp949', 'euc-kr', 'latin-1']
            for enc in encodings_to_try:
                if not enc: continue
                try:
                    html_content = processed_data.decode(enc, errors='replace')
                    print(f"[DEBUG] Step 2 SUCCESS: Decoded with {enc}", flush=True)
                    break
                except Exception:
                    continue
            
            if not html_content:
                html_content = processed_data.decode('utf-8', errors='replace')
                print(f"[DEBUG] Step 2: Last resort decode with utf-8(replace)", flush=True)

        # 4. Parsing
        print(f"[DEBUG] Step 3: Parsing with BeautifulSoup...", flush=True)
        soup = BeautifulSoup(html_content, 'html.parser')
        
        # Clean soup
        for el in soup(["script", "style", "meta", "link", "noscript", "header", "footer", "nav", "iframe"]):
            el.decompose()
            
        # Target content
        article_text = ""
        selectors = [
            'article', '[role="article"]', '.article-content', '.article-body',
            '#article-body', '#articleBody', '.news-content', '#newsct_article',
            'main', '.content', '.post-content', '.entry-content'
        ]
        
        for selector in selectors:
            found = soup.select_one(selector)
            if found:
                text = found.get_text(separator=' ', strip=True)
                if len(text) > 200:
                    article_text = text
                    print(f"[DEBUG] Found content with selector: {selector}", flush=True)
                    break
                    
        if not article_text or len(article_text) < 200:
            body = soup.find('body')
            if body:
                article_text = body.get_text(separator=' ', strip=True)
                print(f"[DEBUG] Using body text as fallback", flush=True)
        
        print(f"[DEBUG] ===== scrape_article SUCCESS: {len(article_text)} chars =====", flush=True)
        return article_text, None
        
    except urllib.error.HTTPError as e:
        msg = f"HTTP Error {e.code}: {e.reason}"
        print(f"[DEBUG] {msg}", flush=True)
        return None, f"웹사이트 접근 실패: {msg}"
    except urllib.error.URLError as e:
        msg = f"URL Error: {e.reason}"
        print(f"[DEBUG] {msg}", flush=True)
        return None, f"URL 오류: {msg}"
    except Exception as e:
        import traceback
        err_msg = str(e)
        print(f"[DEBUG] CRITICAL ERROR in scrape_article: {err_msg}", flush=True)
        print(traceback.format_exc(), flush=True)
        return None, f"크롤링 중 오류 발생: {err_msg}"

def is_korean_text(text):
    """Check if text contains Korean characters"""
    korean_pattern = re.compile(r'[가-힣]')
    return bool(korean_pattern.search(text))

def calculate_arduino_value(progressive_score, conservative_score, total_score):
    """
    Calculate Arduino scale value:
    - Most progressive: 1
    - Neutral: 500
    - Most conservative: 1000
    """
    if total_score == 0:
        return 500  # Neutral when no keywords matched
    
    # Calculate ratio
    progressive_ratio = progressive_score / total_score
    conservative_ratio = conservative_score / total_score
    
    # Map to scale: 1 (progressive) to 1000 (conservative)
    # Progressive: 1-500, Conservative: 500-1000
    if progressive_score > conservative_score:
        # Progressive range: 1 to 500
        # More progressive = closer to 1
        value = 1 + (progressive_ratio * 499)
    elif conservative_score > progressive_score:
        # Conservative range: 500 to 1000
        # More conservative = closer to 1000
        value = 500 + (conservative_ratio * 500)
    else:
        # Equal scores = neutral
        value = 500
    
    return int(round(value))

def split_keywords(cell_value):
    """Split a cell value into multiple keywords if it contains delimiters"""
    if not cell_value:
        return []
    
    text = str(cell_value).strip()
    if not text:
        return []
    
    # Common delimiters: comma, semicolon, pipe, newline, tab, multiple spaces
    # Split by common delimiters
    delimiters = [',', ';', '|', '\n', '\t', '  ']  # double space for multiple spaces
    
    keywords = [text]  # Start with the whole text
    
    # Split by each delimiter
    for delimiter in delimiters:
        new_keywords = []
        for kw in keywords:
            parts = kw.split(delimiter)
            new_keywords.extend([part.strip() for part in parts if part.strip()])
        keywords = new_keywords
    
    # Filter out empty strings and return unique keywords
    return [kw for kw in keywords if kw]

def analyze_political_tendency(text, progressive_keywords, conservative_keywords):
    """Analyze text and determine political tendency"""
    if not text:
        return {
            'tendency': 'Unknown',
            'progressive_score': 0,
            'conservative_score': 0,
            'confidence': 0.0,
            'matched_keywords': {
                'progressive': [],
                'conservative': []
            }
        }
    
    # Check if text contains Korean characters
    is_korean = is_korean_text(text)
    
    # For Korean text, don't convert to lowercase (Korean doesn't have case)
    # For non-Korean text, use lowercase
    if is_korean:
        text_normalized = text
    else:
        text_normalized = text.lower()
    
    # Count keyword matches
    progressive_matches = []
    conservative_matches = []
    
    for keyword in progressive_keywords:
        if not keyword or not keyword.strip():
            continue
            
        keyword_normalized = keyword if is_korean_text(keyword) else keyword.lower()
        
        # For Korean text, use simple substring matching (no word boundaries)
        # For non-Korean text, use word boundaries
        if is_korean_text(keyword) or is_korean:
            # Simple substring matching for Korean
            count = text_normalized.count(keyword_normalized)
            if count > 0:
                progressive_matches.extend([keyword] * count)
        else:
            # Word boundary matching for English
            pattern = r'\b' + re.escape(keyword_normalized) + r'\b'
            matches = re.findall(pattern, text_normalized)
            if matches:
                progressive_matches.extend([keyword] * len(matches))
    
    for keyword in conservative_keywords:
        if not keyword or not keyword.strip():
            continue
            
        keyword_normalized = keyword if is_korean_text(keyword) else keyword.lower()
        
        # For Korean text, use simple substring matching (no word boundaries)
        # For non-Korean text, use word boundaries
        if is_korean_text(keyword) or is_korean:
            # Simple substring matching for Korean
            count = text_normalized.count(keyword_normalized)
            if count > 0:
                conservative_matches.extend([keyword] * count)
        else:
            # Word boundary matching for English
            pattern = r'\b' + re.escape(keyword_normalized) + r'\b'
            matches = re.findall(pattern, text_normalized)
            if matches:
                conservative_matches.extend([keyword] * len(matches))
    
    progressive_score = len(progressive_matches)
    conservative_score = len(conservative_matches)
    total_score = progressive_score + conservative_score
    
    # Determine tendency
    if total_score == 0:
        tendency = 'Neutral'
        confidence = 0.0
    elif progressive_score > conservative_score:
        tendency = 'Progressive'
        confidence = progressive_score / total_score if total_score > 0 else 0.0
    elif conservative_score > progressive_score:
        tendency = 'Conservative'
        confidence = conservative_score / total_score if total_score > 0 else 0.0
    else:
        tendency = 'Mixed'
        confidence = 0.5
    
    # Calculate Arduino scale value (1 = most progressive, 500 = neutral, 1000 = most conservative)
    arduino_value = calculate_arduino_value(progressive_score, conservative_score, total_score)
    
    return {
        'tendency': tendency,
        'progressive_score': progressive_score,
        'conservative_score': conservative_score,
        'confidence': round(confidence * 100, 2),
        'arduino_value': arduino_value,
        'matched_keywords': {
            'progressive': list(set(progressive_matches)),
            'conservative': list(set(conservative_matches))
        }
    }

# Global error handler for unhandled exceptions
@app.errorhandler(500)
def internal_error(error):
    """Handle 500 Internal Server Error"""
    import traceback
    error_trace = traceback.format_exc()
    print(f"[DEBUG] ===== 500 Internal Server Error Handler Called =====")
    print(f"[DEBUG] Error: {error}")
    print(f"[DEBUG] Traceback:")
    print(error_trace)
    return jsonify({'error': 'Internal server error occurred. Check server logs for details.'}), 500

@app.errorhandler(Exception)
def handle_exception(e):
    """Handle all unhandled exceptions"""
    import traceback
    error_trace = traceback.format_exc()
    print(f"[DEBUG] ===== Global Exception Handler Called =====")
    print(f"[DEBUG] Exception type: {type(e).__name__}")
    print(f"[DEBUG] Exception message: {e}")
    print(f"[DEBUG] Full traceback:")
    print(error_trace)
    return jsonify({'error': f'An error occurred: {str(e)}'}), 500

@app.route('/')
def index():
    """Main page"""
    try:
        return render_template('index.html')
    except Exception as e:
        print(f"Error rendering index.html: {e}")
        import traceback
        traceback.print_exc()
        return f"Error loading page: {str(e)}", 500

@app.route('/analyze', methods=['POST'])
def analyze():
    """Analyze a news URL or pasted text"""
    import sys
    import logging
    logger = logging.getLogger(__name__)
    
    # Force output immediately
    sys.stdout.flush()
    sys.stderr.flush()
    
    print("[DEBUG] ===== /analyze endpoint called =====", file=sys.stderr, flush=True)
    print("[DEBUG] ===== /analyze endpoint called =====", flush=True)
    logger.error("[DEBUG] ===== /analyze endpoint called =====")
    app.logger.error("[DEBUG] ===== /analyze endpoint called =====")
    
    try:
        print("[DEBUG] Step A1: Getting request data", flush=True)
        logger.error("[DEBUG] Step A1: Getting request data")
        data = request.get_json()
        if not data:
            print("[DEBUG] Step A1 FAILED: No data provided")
            return jsonify({'error': 'No data provided'}), 400
        
        url = data.get('url', '').strip()
        text = data.get('text', '').strip()
        print(f"[DEBUG] Step A1 SUCCESS: url='{url[:50]}...' if url else None, text length={len(text) if text else 0}")
        
        # Load keywords
        print("[DEBUG] Step A2: Loading keywords")
        try:
            progressive_keywords, conservative_keywords = load_keywords()
            print(f"[DEBUG] Step A2 SUCCESS: Loaded {len(progressive_keywords)} progressive, {len(conservative_keywords)} conservative keywords")
        except Exception as e:
            print(f"[DEBUG] Step A2 FAILED: {type(e).__name__}: {e}")
            import traceback
            traceback.print_exc()
            return jsonify({'error': f'Error loading keywords: {str(e)}'}), 500
        
        article_text = None
        source = None
        
        # Check if text is provided (paste mode)
        if text:
            print("[DEBUG] Step A3: Using pasted text")
            article_text = text
            source = 'pasted_text'
            print(f"[DEBUG] Step A3 SUCCESS: Using pasted text, length={len(article_text)}")
        # Otherwise, try to scrape from URL
        elif url:
            print(f"[DEBUG] Step A3: Scraping URL: {url}")
            # Validate URL
            try:
                parsed = urlparse(url)
                if not parsed.scheme or not parsed.netloc:
                    print("[DEBUG] Step A3 FAILED: Invalid URL format")
                    return jsonify({'error': 'Invalid URL format'}), 400
                print("[DEBUG] Step A3.1: URL validation passed")
            except Exception as e:
                print(f"[DEBUG] Step A3.1 FAILED: {type(e).__name__}: {e}")
                return jsonify({'error': f'Invalid URL format: {str(e)}'}), 400
            
            # Scrape article
            print("[DEBUG] Step A3.2: Calling scrape_article()")
            try:
                article_text, error = scrape_article(url)
                print(f"[DEBUG] Step A3.2: scrape_article returned, error={error is not None}")
                if error:
                    print(f"[DEBUG] Step A3.2 FAILED: {error}")
                    # Check if it's a decompression error - return 400 instead of 500
                    if "압축" in error or "decompressing" in error.lower() or "zlib" in error.lower() or "incorrect header" in error.lower():
                        print("[DEBUG] Returning 400 for decompression error")
                        return jsonify({'error': error}), 400
                    # For other errors, return 500
                    print("[DEBUG] Returning 500 for other error")
                    return jsonify({'error': error}), 500
                print(f"[DEBUG] Step A3.2 SUCCESS: Got article_text, length={len(article_text) if article_text else 0}")
            except Exception as e:
                error_str = str(e)
                print(f"[DEBUG] Step A3.2 EXCEPTION: {type(e).__name__}: {error_str}")
                import traceback
                print("[DEBUG] Full traceback:")
                traceback.print_exc()
                # Check if it's a decompression error
                if "decompressing" in error_str.lower() or "zlib" in error_str.lower() or "incorrect header" in error_str.lower() or "-3" in error_str:
                    print("[DEBUG] Returning 400 for decompression error (exception)")
                    return jsonify({'error': f'웹사이트 데이터 처리 중 오류가 발생했습니다: {error_str}'}), 400
                print("[DEBUG] Returning 500 for other error (exception)")
                return jsonify({'error': f'Error scraping article: {error_str}'}), 500
            source = url
        else:
            print("[DEBUG] Step A3 FAILED: Neither URL nor text provided")
            return jsonify({'error': 'Either URL or text content is required'}), 400
        
        # Analyze political tendency
        print("[DEBUG] Step A4: Analyzing political tendency")
        try:
            result = analyze_political_tendency(article_text, progressive_keywords, conservative_keywords)
            result['source'] = source
            result['article_length'] = len(article_text) if article_text else 0
            result['article_text'] = article_text if article_text else ''
            print(f"[DEBUG] Step A4 SUCCESS: Tendency={result.get('tendency')}, Confidence={result.get('confidence')}")
        except Exception as e:
            print(f"[DEBUG] Step A4 FAILED: {type(e).__name__}: {e}")
            import traceback
            traceback.print_exc()
            return jsonify({'error': f'Error analyzing article: {str(e)}'}), 500
        
        # Send Arduino value via serial (with error handling)
        print("[DEBUG] Step A5: Sending to Arduino")
        if SERIAL_SUPPORT:
            try:
                send_to_arduino(result['arduino_value'])
                print(f"[DEBUG] Step A5 SUCCESS: Sent {result['arduino_value']} to Arduino")
            except Exception as e:
                print(f"[DEBUG] Step A5 WARNING: Failed to send to Arduino: {e}")
                # Continue even if Arduino send fails
        
        print("[DEBUG] ===== /analyze endpoint SUCCESS =====")
        return jsonify(result)
    except Exception as e:
        print(f"[DEBUG] ===== /analyze endpoint EXCEPTION (outer catch) =====")
        print(f"[DEBUG] Exception type: {type(e).__name__}")
        print(f"[DEBUG] Exception message: {e}")
        import traceback
        print("[DEBUG] Full traceback:")
        traceback.print_exc()
        return jsonify({'error': f'Unexpected error: {str(e)}'}), 500

@app.route('/health', methods=['GET'])
def health():
    """Health check endpoint"""
    return jsonify({'status': 'healthy'})

@app.route('/keywords/count', methods=['GET'])
def get_keyword_counts():
    """Get current keyword counts"""
    progressive_keywords, conservative_keywords = load_keywords()
    return jsonify({
        'progressive_count': len(progressive_keywords),
        'conservative_count': len(conservative_keywords)
    })

@app.route('/keywords/upload', methods=['POST'])
def upload_keywords():
    """Upload keywords file (Excel or JSON)"""
    try:
        # Debug: Check what's in the request
        print(f"DEBUG: Request files: {list(request.files.keys())}")
        print(f"DEBUG: Request form: {dict(request.form)}")
        
        if 'file' not in request.files:
            print("DEBUG: No 'file' key in request.files")
            return jsonify({'error': 'No file provided in request'}), 400
        
        file = request.files['file']
        keyword_type = request.form.get('type', '').strip()
        
        print(f"DEBUG: File object: {file}")
        print(f"DEBUG: File filename: {file.filename if file else 'None'}")
        print(f"DEBUG: Keyword type: '{keyword_type}'")
        
        if not file:
            return jsonify({'error': 'File object is None'}), 400
        
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        if not keyword_type:
            return jsonify({'error': 'Keyword type not provided'}), 400
        
        if keyword_type not in ['progressive', 'conservative']:
            return jsonify({'error': f'Invalid keyword type: {keyword_type}. Must be "progressive" or "conservative"'}), 400
        
        # Get original filename before secure_filename
        original_filename = file.filename
        print(f"DEBUG: Original filename: {original_filename}")
        
        # Extract extension from original filename first
        if '.' in original_filename:
            file_ext = original_filename.rsplit('.', 1)[1].lower()
        else:
            file_ext = ''
        
        print(f"DEBUG: Detected file extension: '{file_ext}'")
        
        if not file_ext:
            return jsonify({'error': f'File has no extension. Original filename: {original_filename}'}), 400
        
        if file_ext not in ['xlsx', 'json']:
            return jsonify({'error': f'Invalid file type: .{file_ext}. Only .xlsx and .json files are supported.'}), 400
        
        # For Korean filenames, preserve the original filename better
        # secure_filename might strip Korean characters, so we'll use a safer approach
        if is_korean_text(original_filename):
            # Keep original filename but sanitize path separators
            filename = original_filename.replace('/', '_').replace('\\', '_')
        else:
            filename = secure_filename(file.filename)
        print(f"DEBUG: Secure filename: {filename}")
        
        # Process the file - save to executable directory when running as executable
        if getattr(sys, 'frozen', False):
            base_dir = os.path.dirname(sys.executable)
        else:
            base_dir = os.path.dirname(__file__)
        
        # Load existing keywords first (to stack/add to them)
        json_path = os.path.join(base_dir, f'{keyword_type}.json')
        existing_keywords = []
        if os.path.exists(json_path):
            try:
                with open(json_path, 'r', encoding='utf-8') as f:
                    existing_keywords = json.load(f)
                    if not isinstance(existing_keywords, list):
                        existing_keywords = []
            except Exception as e:
                print(f"Warning: Could not load existing keywords: {e}")
                existing_keywords = []
        
        new_keywords = []
        
        if file_ext == 'json':
            # Handle JSON file
            keywords_raw = json.load(file)
            if not isinstance(keywords_raw, list):
                return jsonify({'error': 'JSON file must contain an array of keywords'}), 400
            
            # Split keywords if they contain delimiters
            for item in keywords_raw:
                if isinstance(item, str):
                    # Split if it contains multiple keywords
                    split_kws = split_keywords(item)
                    new_keywords.extend(split_kws)
                else:
                    new_keywords.append(str(item))
            
            new_keywords = [kw for kw in new_keywords if kw]
            
        elif file_ext == 'xlsx':
            # Handle Excel file
            if not EXCEL_SUPPORT:
                return jsonify({'error': 'Excel support not available. Install openpyxl.'}), 500
            
            # Save uploaded file temporarily (handle Korean filenames)
            temp_path = os.path.join(base_dir, f'temp_{keyword_type}.xlsx')
            # Ensure the file is saved with proper encoding
            file.save(temp_path)
            
            # Read Excel file
            wb = load_workbook(temp_path, read_only=True)
            ws = wb.active
            
            # Read from all columns, not just the first one
            max_col = ws.max_column
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=max_col, values_only=True):
                for cell_value in row:
                    if cell_value:
                        # Split cell value into multiple keywords if it contains delimiters
                        cell_keywords = split_keywords(cell_value)
                        for kw in cell_keywords:
                            # Skip header rows (check both English and Korean)
                            kw_lower = kw.lower()
                            header_keywords = ['keyword', 'keywords', 'term', 'terms', keyword_type, 
                                              '키워드', '단어', '용어', '진보', '보수', '진보진영', '보수진영']
                            if kw_lower not in header_keywords and kw not in header_keywords:
                                new_keywords.append(kw)
            
            new_keywords = [kw for kw in new_keywords if kw]
            wb.close()
            
            # Save as JSON (also keep Excel if user wants)
            excel_path = os.path.join(base_dir, f'{keyword_type}.xlsx')
            shutil.move(temp_path, excel_path)
        
        # Combine existing and new keywords, removing duplicates
        # For Korean keywords, use exact match; for others, use case-insensitive
        combined_keywords = list(existing_keywords)  # Start with existing
        
        for new_kw in new_keywords:
            # Check if keyword already exists
            is_duplicate = False
            for existing_kw in combined_keywords:
                if is_korean_text(new_kw) or is_korean_text(existing_kw):
                    # Exact match for Korean
                    if new_kw == existing_kw:
                        is_duplicate = True
                        break
                else:
                    # Case-insensitive match for non-Korean
                    if new_kw.lower() == existing_kw.lower():
                        is_duplicate = True
                        break
            
            if not is_duplicate:
                combined_keywords.append(new_kw)
        
        # Save combined keywords to JSON file
        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump(combined_keywords, f, indent=2, ensure_ascii=False)
        
        added_count = len(combined_keywords) - len(existing_keywords)
        
        return jsonify({
            'success': True,
            'count': len(combined_keywords),
            'added': added_count,
            'existing': len(existing_keywords),
            'message': f'기존 {len(existing_keywords)}개 키워드에 {added_count}개 추가됨 (총 {len(combined_keywords)}개)'
        })
        
    except json.JSONDecodeError as e:
        print(f"DEBUG: JSON decode error: {str(e)}")
        return jsonify({'error': f'Invalid JSON file format: {str(e)}'}), 400
    except Exception as e:
        print(f"DEBUG: Exception in upload_keywords: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Error processing file: {str(e)}'}), 500

@app.route('/keywords/download/<keyword_type>', methods=['GET'])
def download_keywords(keyword_type):
    """Download keywords as JSON"""
    if keyword_type not in ['progressive', 'conservative']:
        return jsonify({'error': 'Invalid keyword type'}), 400
    
    progressive_keywords, conservative_keywords = load_keywords()
    
    keywords = progressive_keywords if keyword_type == 'progressive' else conservative_keywords
    
    return jsonify({'keywords': keywords})

@app.route('/arduino/connect', methods=['POST'])
def connect_arduino():
    """Manually connect to Arduino via serial"""
    try:
        data = request.get_json() or {}
        port = data.get('port', '').strip()
        baudrate = data.get('baudrate', 9600)
        
        if not SERIAL_SUPPORT:
            return jsonify({'error': 'pyserial not installed'}), 500
        
        if init_arduino_serial(port if port else None, baudrate):
            return jsonify({
                'success': True,
                'port': serial_port,
                'message': f'Arduino connected on {serial_port}'
            })
        else:
            return jsonify({'error': 'Failed to connect to Arduino. Check if the port is correct and Arduino is connected.'}), 500
    except Exception as e:
        print(f"Error connecting to Arduino: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Error connecting to Arduino: {str(e)}'}), 500

@app.route('/arduino/disconnect', methods=['POST'])
def disconnect_arduino():
    """Disconnect Arduino serial connection"""
    global arduino_serial, serial_port
    
    try:
        port = None
        if arduino_serial is not None:
            try:
                if arduino_serial.is_open:
                    arduino_serial.close()
            except (AttributeError, ValueError, OSError) as e:
                print(f"Error closing serial port: {e}")
            finally:
                port = serial_port
                arduino_serial = None
                serial_port = None
        
        if port:
            return jsonify({'success': True, 'message': f'Disconnected from {port}'})
        else:
            return jsonify({'error': 'No active connection'}), 400
    except Exception as e:
        print(f"Error disconnecting Arduino: {e}")
        arduino_serial = None
        serial_port = None
        return jsonify({'error': f'Error disconnecting: {str(e)}'}), 500

@app.route('/arduino/status', methods=['GET'])
def arduino_status():
    """Get Arduino connection status"""
    global arduino_serial, serial_port
    
    if not SERIAL_SUPPORT:
        return jsonify({
            'connected': False,
            'message': 'pyserial not installed',
            'available_ports': []
        })
    
    # Check connection status with error handling
    try:
        if arduino_serial is not None:
            try:
                is_open = arduino_serial.is_open
                if is_open:
                    return jsonify({
                        'connected': True,
                        'port': serial_port,
                        'message': f'Connected on {serial_port}'
                    })
            except (AttributeError, ValueError, OSError) as e:
                # Serial port is in invalid state, reset it
                print(f"Serial port in invalid state: {e}")
                arduino_serial = None
                serial_port = None
    except Exception as e:
        print(f"Error checking Arduino status: {e}")
        arduino_serial = None
        serial_port = None
    
    # List available ports
    ports = []
    if SERIAL_SUPPORT:
        try:
            available_ports = serial.tools.list_ports.comports()
            ports = [{'device': p.device, 'description': p.description} for p in available_ports]
        except Exception as e:
            print(f"Error listing ports: {e}")
    
    return jsonify({
        'connected': False,
        'message': 'Not connected',
        'available_ports': ports
    })

@app.route('/arduino/value', methods=['GET'])
def get_arduino_value():
    """
    Get Arduino scale value for the last analyzed article.
    Returns a simple integer value:
    - 1 = Most progressive
    - 500 = Neutral
    - 1000 = Most conservative
    """
    # For simplicity, return a default neutral value
    # In a real implementation, you might want to store the last analysis result
    return str(500)

@app.route('/arduino/analyze', methods=['POST'])
def analyze_for_arduino():
    """
    Analyze article and return Arduino value directly.
    Accepts URL or text, returns just the numeric value for Arduino.
    """
    data = request.get_json()
    url = data.get('url', '').strip()
    text = data.get('text', '').strip()
    
    # Load keywords
    progressive_keywords, conservative_keywords = load_keywords()
    
    article_text = None
    
    # Check if text is provided (paste mode)
    if text:
        article_text = text
    # Otherwise, try to scrape from URL
    elif url:
        # Validate URL
        try:
            parsed = urlparse(url)
            if not parsed.scheme or not parsed.netloc:
                return jsonify({'error': 'Invalid URL format'}), 400
        except Exception:
            return jsonify({'error': 'Invalid URL format'}), 400
        
        # Scrape article
        article_text, error = scrape_article(url)
        if error:
            return jsonify({'error': error}), 500
    else:
        return jsonify({'error': 'Either URL or text content is required'}), 400
    
    # Analyze political tendency
    result = analyze_political_tendency(article_text, progressive_keywords, conservative_keywords)
    
    # Return just the Arduino value as plain text (easier for Arduino to parse)
    return str(result['arduino_value']), 200, {'Content-Type': 'text/plain; charset=utf-8'}

if __name__ == '__main__':
    import socket
    
    # Load keywords on startup to validate files
    try:
        progressive_keywords, conservative_keywords = load_keywords()
        print(f"Loaded {len(progressive_keywords)} progressive keywords")
        print(f"Loaded {len(conservative_keywords)} conservative keywords")
    except Exception as e:
        print(f"ERROR loading keywords: {e}")
        import traceback
        traceback.print_exc()
        progressive_keywords, conservative_keywords = [], []
        
    print(f"==================================================")
    print(f"Political Tendency Analyzer Server - v{VERSION}")
    print(f"==================================================")
    
    # Get local IP address
    def get_local_ip():
        """Get the local IP address"""
        try:
            # Method 1: Connect to a remote address to determine local IP
            s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
            s.connect(("8.8.8.8", 80))
            ip = s.getsockname()[0]
            s.close()
            return ip
        except Exception:
            try:
                # Method 2: Get hostname and resolve
                hostname = socket.gethostname()
                ip = socket.gethostbyname(hostname)
                # Filter out localhost
                if ip != "127.0.0.1":
                    return ip
            except Exception:
                pass
            return None
    
    def get_all_ips():
        """Get all network IP addresses"""
        ips = []
        try:
            hostname = socket.gethostname()
            # Get all IP addresses
            for addr_info in socket.getaddrinfo(hostname, None):
                ip = addr_info[4][0]
                if ip not in ['127.0.0.1', '::1'] and ip not in ips:
                    # Filter IPv4 addresses
                    if '.' in ip:
                        ips.append(ip)
        except Exception:
            pass
        return ips
    
    local_ip = get_local_ip()
    all_ips = get_all_ips()
    port = 5000
    
    print("\n" + "="*50)
    print("Political Tendency Analyzer Server")
    print("="*50)
    print(f"Server starting on:")
    print(f"  Local:   http://127.0.0.1:{port}")
    
    if local_ip:
        print(f"  Network: http://{local_ip}:{port}")
    else:
        print("  Network: Could not detect IP automatically")
    
    if all_ips:
        print("\n  Detected IP addresses:")
        for ip in all_ips:
            print(f"    - http://{ip}:{port}")
    
    print("\nTo access from other devices on the same network:")
    if local_ip:
        print(f"  Open browser and go to: http://{local_ip}:{port}")
    elif all_ips:
        print(f"  Try one of these IPs: {', '.join([f'http://{ip}:{port}' for ip in all_ips])}")
    else:
        print("  Find your IP address manually:")
        print("    Windows: ipconfig (look for IPv4 Address)")
        print("    Mac/Linux: ifconfig or ip addr")
    
    print("\nTroubleshooting:")
    print("  1. Make sure Windows Firewall allows Python through")
    print("  2. Ensure both devices are on the same Wi-Fi/network")
    print("  3. Try accessing from the server machine first: http://127.0.0.1:5000")
    print("="*50 + "\n")
    
    try:
        print(f"\nStarting server on 0.0.0.0:{port}...")
        print("Server is running. Press Ctrl+C to stop.\n")
        # Use 0.0.0.0 to bind to all network interfaces (not just localhost)
        # This allows access from other devices on the network
        try:
            app.run(debug=False, host='0.0.0.0', port=port, threaded=True, use_reloader=False)
        finally:
            # Ensure Arduino is disconnected when server stops
            cleanup_arduino()
    except OSError as e:
        if "Address already in use" in str(e) or "10048" in str(e):
            print(f"\nERROR: Port {port} is already in use!")
            print("Please close the application using port 5000 or change the port number.")
            print("\nTrying to start on port 5001 instead...")
            try:
                try:
                    app.run(debug=False, host='0.0.0.0', port=5001, threaded=True, use_reloader=False)
                finally:
                    # Ensure Arduino is disconnected when server stops
                    cleanup_arduino()
            except Exception as e2:
                print(f"ERROR: Could not start on port 5001 either: {e2}")
                cleanup_arduino()
        else:
            print(f"\nERROR: {e}")
            import traceback
            traceback.print_exc()

